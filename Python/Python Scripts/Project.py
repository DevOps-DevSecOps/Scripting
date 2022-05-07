       
from openpyxl import Workbook
import openpyxl
import pymysql
import smtplib

con = pymysql.connect(host='localhost',port=3306,user='root',password='Admin@123',db='atk')
cursor = con.cursor()

book = Workbook()
sheet = book.active
sheet['A1'] = 'aruntheja9394@gmail.com'
sheet['A2'] = 'aruntheja9493@gmail.com'
sheet['A3'] = 'aruntheja9494@gmail.com'
sheet['A4'] = 'tspavankumart@gmail.com'
book.save("sample.xlsx")

book = openpyxl.load_workbook('sample.xlsx')
sheet = book.active
row_count = sheet.max_row
column_count = sheet.max_column
for i in range(1,row_count+1):
	a = sheet['A'+str(i)]
	sql = 'insert into list(Emails) values ('+a.value+');'
	cursor.execute(sql)
	con.commit()

sql = 'select Emails from list;'
cursor.execute(sql)
rows = cursor.fetchall()
for item in rows:
	print(item)

	file = open('text','r')
	print(file.read)

	username = 'aruntheja94@gmail.com'
	password = 'atk@1994'
	print('login successfully')

	server = smtplib.SMTP('smtp.gmail.com:587')
	server.starttls()
	server.login(username,password)
	server.sendmail(username,item, file.read())
	server.quit() 

sql='TRUNCATE list'
cursor.execute(sql)
con.commit()

cursor.close()
con.close()